import hashlib
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction as db_transaction
from openpyxl import load_workbook

from apps.assets.models import Asset, Account
from apps.transactions.models import Transaction, Dividend, Interest


def _to_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _to_decimal(val, default=None):
    if val is None:
        return default
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return default


def _hash(*parts):
    raw = "|".join(str(p) for p in parts)
    return hashlib.sha256(raw.encode()).hexdigest()


ORDER_MAP = {
    "Compra": Transaction.TransactionType.BUY,
    "Venta": Transaction.TransactionType.SELL,
    "Vendido": Transaction.TransactionType.SELL,
    "Regalo": Transaction.TransactionType.GIFT,
}


def _clean_headers(ws):
    """Read headers from row 1, truncating at the first None to ignore summary columns."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        end = headers.index(None)
        return headers[:end]
    except ValueError:
        return headers


def import_xlsx(file, dry_run=False):
    wb = load_workbook(file, data_only=True)
    result = {
        "inserted": {"transactions": 0, "dividends": 0, "interests": 0, "assets": 0},
        "skipped_duplicates": {"transactions": 0, "dividends": 0, "interests": 0},
        "errors": [],
    }

    default_account, _ = Account.objects.get_or_create(
        name="Trade Republic",
        defaults={"type": Account.AccountType.INVERSION},
    )

    asset_cache = {}
    for asset in Asset.objects.all():
        if asset.ticker:
            asset_cache[asset.ticker] = asset
        asset_cache[asset.name] = asset

    def get_or_create_asset(name, ticker=None, current_price=None):
        key = ticker or name
        if key in asset_cache:
            return asset_cache[key]
        if name in asset_cache:
            return asset_cache[name]
        asset = Asset.objects.create(
            name=name,
            ticker=ticker if ticker else None,
            current_price=current_price,
        )
        result["inserted"]["assets"] += 1
        asset_cache[name] = asset
        if ticker:
            asset_cache[ticker] = asset
        return asset

    with db_transaction.atomic():
        # --- Acciones sheet ---
        if "Acciones" in wb.sheetnames:
            ws = wb["Acciones"]
            headers = _clean_headers(ws)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_data = dict(zip(headers, row))
                dt = _to_date(row_data.get("F.Valor"))
                order = row_data.get("Orden")
                entity = row_data.get("Entidad")
                ticker = row_data.get("Ticker")

                if dt is None or order is None or entity is None:
                    if any(v is not None for v in row):
                        result["errors"].append({
                            "sheet": "Acciones", "row": row_idx,
                            "column": "F.Valor/Orden/Entidad",
                            "message": "Missing required field",
                        })
                    continue

                tx_type = ORDER_MAP.get(order)
                if tx_type is None:
                    result["errors"].append({
                        "sheet": "Acciones", "row": row_idx,
                        "column": "Orden",
                        "message": f"Unknown order type: {order}",
                    })
                    continue

                quantity = _to_decimal(row_data.get("Cantidad Incremental"), Decimal("0"))
                price = _to_decimal(row_data.get("Precio"))
                commission = _to_decimal(row_data.get("Comision"), Decimal("0"))
                tax = _to_decimal(row_data.get("Impuesto"), Decimal("0"))

                h = _hash("TX", dt, order, entity, ticker, quantity, price, commission)

                if Transaction.objects.filter(import_hash=h).exists():
                    result["skipped_duplicates"]["transactions"] += 1
                    continue

                current_price_val = _to_decimal(row_data.get("Cotización actual"))
                asset = get_or_create_asset(entity, ticker, current_price_val)

                Transaction.objects.create(
                    date=dt,
                    type=tx_type,
                    asset=asset,
                    account=default_account,
                    quantity=abs(quantity),
                    price=price,
                    commission=commission,
                    tax=tax,
                    import_hash=h,
                )
                result["inserted"]["transactions"] += 1

        # --- Dividendos sheet ---
        if "Dividendos" in wb.sheetnames:
            ws = wb["Dividendos"]
            headers = _clean_headers(ws)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_data = dict(zip(headers, row))
                dt = _to_date(row_data.get("F.Valor"))
                entity = row_data.get("Entidad")

                if dt is None or entity is None:
                    if any(v is not None for v in row):
                        result["errors"].append({
                            "sheet": "Dividendos", "row": row_idx,
                            "column": "F.Valor/Entidad",
                            "message": "Missing required field",
                        })
                    continue

                gross = _to_decimal(row_data.get("Precio"), Decimal("0"))
                tax = _to_decimal(row_data.get("Impuestos"), Decimal("0"))
                net = _to_decimal(row_data.get("Total"), Decimal("0"))
                shares = _to_decimal(row_data.get("Acciones"))
                withholding = _to_decimal(row_data.get("% Retención"))

                h = _hash("DIV", dt, entity, gross, tax, net)

                if Dividend.objects.filter(import_hash=h).exists():
                    result["skipped_duplicates"]["dividends"] += 1
                    continue

                asset = get_or_create_asset(entity)

                Dividend.objects.create(
                    date=dt,
                    asset=asset,
                    shares=shares,
                    gross=gross,
                    tax=tax,
                    net=net,
                    withholding_rate=withholding,
                    import_hash=h,
                )
                result["inserted"]["dividends"] += 1

        # --- Intereses sheet ---
        if "Intereses" in wb.sheetnames:
            ws = wb["Intereses"]
            headers = _clean_headers(ws)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_data = dict(zip(headers, row))
                dt = _to_date(row_data.get("F.Valor"))
                entity = row_data.get("Entidad")

                if dt is None or entity is None:
                    if any(v is not None for v in row):
                        result["errors"].append({
                            "sheet": "Intereses", "row": row_idx,
                            "column": "F.Valor/Entidad",
                            "message": "Missing required field",
                        })
                    continue

                gross = _to_decimal(row_data.get("Interes Bruto"), Decimal("0"))
                net = _to_decimal(row_data.get("Neto"), Decimal("0"))
                balance = _to_decimal(row_data.get("Saldo"))
                annual_rate = _to_decimal(row_data.get("% Interes anual"))

                h = _hash("INT", dt, entity, gross, net)

                if Interest.objects.filter(import_hash=h).exists():
                    result["skipped_duplicates"]["interests"] += 1
                    continue

                account, _ = Account.objects.get_or_create(
                    name=entity,
                    defaults={"type": Account.AccountType.AHORRO},
                )

                Interest.objects.create(
                    date=dt,
                    account=account,
                    gross=gross,
                    net=net,
                    balance=balance,
                    annual_rate=annual_rate,
                    import_hash=h,
                )
                result["inserted"]["interests"] += 1

        # --- Entidades sheet (update current prices) ---
        if "Entidades" in wb.sheetnames:
            ws = wb["Entidades"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                name = row_data.get("Entidad")
                ticker = row_data.get("Ticker")
                price = _to_decimal(row_data.get("Cotización"))
                if name and price is not None:
                    asset = get_or_create_asset(name, ticker, price)
                    if asset.current_price != price:
                        asset.current_price = price
                        asset.save(update_fields=["current_price"])

        if dry_run:
            db_transaction.set_rollback(True)

    return result
